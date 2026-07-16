# ---------------------------------------------------------------------------
# api/index.py -- Vercel serverless entrypoint.
#
# This is the Flask app AND the traverse-generation engine in one file
# (merged deliberately -- Vercel's Python runtime is most reliable when a
# function's entrypoint doesn't depend on sibling-module imports). Vercel
# auto-detects the top-level `app` object below and serves it as a
# serverless function at /api/index (routed to /api/* by vercel.json).
#
# Running locally without Vercel: `python api/index.py` still works (see
# the __main__ block at the bottom) for standalone Flask dev/testing.
# ---------------------------------------------------------------------------

"""
traverse_generator.py

Rebuilds the "Form 10" traverse-computation workbook (DTM, TRVS, ABSTRACT, AREA
sheets) for any number of corner points, from a plain JSON payload.

This mirrors the formula patterns of the original Form_10.xls exactly (see
Traverse_Form_Extension_Guide.docx for the derivation) but generates them
programmatically instead of relying on someone hand-editing rows in Excel.

Design assumptions (documented so they're easy to challenge/change):
  1. The opening block is always: 2 backsight control stations + 1 start
     station (the traverse's first point). The closing block mirrors this:
     2 foresight control stations + 1 end station.
  2. "Plot corners" are whichever entered corner points (plus the start/end
     stations, if tagged) the caller flags as `is_plot_corner`. ABSTRACT and
     AREA are built only from that tagged subset, in traverse order.
  3. AREA closes the polygon by repeating the first plot-corner point as the
     last point, so the shoelace-formula area is always for a closed shape.
     (The original file did not always do this consistently -- some jobs
     left stray #REF! rows from an oversized master template. This
     implementation always produces a mathematically closed, error-free
     polygon.)
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    _IMPORT_ERROR = None
except Exception as _e:  # pragma: no cover -- surfaces a readable error instead of a bare crash
    _IMPORT_ERROR = _e

FONT_NAME = "Arial"

# Colors verified against Form_10.xls and Form_6.xls (identical scheme in
# both real files):
#   FF0066CC -- a value that traces directly back to a DTM control station
#               (station name/coords, or a bearing pulled straight from
#               DTM rather than computed via ATAN2), plus the Distance/
#               Meas-Dist/Corr(N)/Corr(E) columns on every row.
#   FF0000FF -- the running angular correction column (E), input or not.
#   FF993300 -- the leg-distance column (O).
#   FF000000 -- everything else that's computed (bearing DMS breakdown,
#               raw deltas, station/coordinate inputs on corner rows).
DTM_LINK_COLOR = "FF0066CC"
CORRECTION_COLOR = "FF0000FF"
LEG_DIST_COLOR = "FF993300"
COMPUTED_COLOR = "FF000000"

INPUT_FONT = Font(name=FONT_NAME, color="0000FF")   # blue = a cell the user typed in
FORMULA_FONT = Font(name=FONT_NAME, color="000000")  # black = computed
LABEL_FONT = Font(name=FONT_NAME, bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
CENTER = Alignment(horizontal="center")

NUMFMT_COORD = "0.000"    # coordinates, distances, corrections
NUMFMT_DIST2 = "0.00"     # a couple of distance-ish columns use 2dp instead of 3
NUMFMT_INT = "0"          # degree/minute/second components


def _set(ws, coord, value, font=FORMULA_FONT, bold=False, color=None,
         center=True, number_format=None):
    cell = ws[coord]
    cell.value = value
    font_color = color or font.color
    cell.font = Font(name=font.name, color=font_color, bold=bold or font.bold)
    if center:
        cell.alignment = CENTER
    if number_format:
        cell.number_format = number_format
    return cell


def _merge(ws, cell_range):
    ws.merge_cells(cell_range)


# --------------------------------------------------------------------------
# TOC sheet -- static informational contents list (no formulas in the
# original either; page numbers are illustrative, not computed).
# --------------------------------------------------------------------------
def build_toc(wb):
    ws = wb.create_sheet("TOC's")
    _set(ws, "D1", "TABLE OF CONTENTS", font=TITLE_FONT, bold=True)
    _set(ws, "B4", "ITEM", bold=True)
    _set(ws, "I4", "PAGE", bold=True)
    items = [
        "Mutation form",
        "Plane page",
        "Table of contents",
        "Job History",
        "Index to computations",
        "Abstract of coordinates",
        "Datum Coordinates and Computations",
        "Field notes",
        "Traverse Computation sheet",
        "Area Computation",
    ]
    for i, item in enumerate(items):
        r = 6 + i
        _set(ws, f"A{r}", i + 1)
        _set(ws, f"B{r}", item + " " + "." * 30)
        _set(ws, f"I{r}", i + 1)
    return ws


# --------------------------------------------------------------------------
# Job History sheet -- basic job metadata (the original left this blank;
# a few structured fields are more useful than an empty page).
# --------------------------------------------------------------------------
def build_job_history(wb, data):
    ws = wb.create_sheet("Job History")
    _set(ws, "A1", "JOB HISTORY", font=TITLE_FONT, bold=True)
    meta = data.get("job_meta", {})
    rows = [
        ("Job number", meta.get("job_number", "")),
        ("Client", meta.get("client", "")),
        ("Surveyor", meta.get("surveyor", "")),
        ("Date", meta.get("date", "")),
        ("Description", meta.get("description", "")),
    ]
    for i, (label, value) in enumerate(rows):
        r = 3 + i
        _set(ws, f"A{r}", label, bold=True)
        _set(ws, f"C{r}", value, font=INPUT_FONT)
    return ws


# --------------------------------------------------------------------------
# INDEX sheet -- the distance-correction / multiplication-factor note.
# The original had 4 near-duplicate draft copies of this block; this
# generates one clean, parameterized copy at the same row numbers as the
# original's last (most complete) block, so Field notes!M6's reference to
# INDEX!F95 keeps working unchanged.
# --------------------------------------------------------------------------
def build_index_sheet(wb, data):
    ws = wb.create_sheet("INDEX")
    idx = data.get("index_notes", {})
    _set(ws, "D80", "INDEX TO COMPUTATIONS", font=TITLE_FONT, bold=True)
    _set(ws, "B83", f'The job was done using the total station {idx.get("instrument", "------------")} observing horizontal distances.')
    _set(ws, "B84", "The following are the standard corrections which were applied to obtain to")
    _set(ws, "B85", "the final distances.")

    _set(ws, "B87", "Height above sea level")
    _set(ws, "E87", "=")
    _set(ws, "F87", idx.get("elevation", 1220), font=INPUT_FONT)
    _set(ws, "G87", f'({idx.get("district", "District")})')

    _set(ws, "B89", "Mean sea level correction")
    _set(ws, "E89", "=")
    _set(ws, "F89", idx.get("msl_correction", -0.1911), font=INPUT_FONT)

    _set(ws, "B91", "Scale factor")
    _set(ws, "E91", "=")
    _set(ws, "F91", idx.get("scale_factor", -0.3737), font=INPUT_FONT)

    _set(ws, "B93", "Combined correction")
    _set(ws, "E93", "=")
    _set(ws, "F93", "=F89+F91")

    _set(ws, "B95", "Multiplication Factor MF")
    _set(ws, "E95", "=")
    _set(ws, "F95", "=(1+F93/1000)")

    _set(ws, "B97", "Note:")
    _set(ws, "B98", "All the other corrections were made automatically done by the total station and so the")
    _set(ws, "B99", "the recorded distances are truly horizontal.")
    return ws



def build_dtm(wb, data):
    ws = wb.create_sheet("DTM")
    _set(ws, "A1", "DATUM COORDINATES AND COMPUTATIONS", font=TITLE_FONT, bold=True)
    _merge(ws, "A1:K1")
    for c, label in zip("ABC", ("Stn", "Northings", "Eastings")):
        _set(ws, f"{c}2", label, bold=True)

    o = data["opening"]
    c_ = data["closing"]

    def _control_row(r, stn):
        _set(ws, f"A{r}", stn["name"], font=INPUT_FONT, color=COMPUTED_COLOR)
        _set(ws, f"B{r}", stn["n"], font=INPUT_FONT, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"C{r}", stn["e"], font=INPUT_FONT, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _merge(ws, f"D{r}:E{r}")

    # Opening control stations (rows 4-6)
    _control_row(4, o["backsight1"])
    _control_row(5, o["backsight2"])
    _control_row(6, o["start"])

    # Closing control stations (rows 7-9).
    # NOTE: A7 must hold the "end" (common) station -- the TRVS closing block
    # references DTM!A7 directly, and DTM!A8/A9 are the two foresight
    # stations whose bearings both resolve onto A7.
    _control_row(7, c_["end"])
    _control_row(8, c_["foresight1"])
    _control_row(9, c_["foresight2"])

    def _echo_row(r, src_r):
        # "Echo" rows just copy a control station's name/coords -- black,
        # matching the original (not the blue "computed bearing" rows).
        _set(ws, f"A{r}", f"=A{src_r}", color=COMPUTED_COLOR)
        _set(ws, f"B{r}", f"=B{src_r}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"C{r}", f"=C{src_r}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)

    def _bearing_row(r, from_r, to_r, headers=None):
        # Row that computes bearing+distance between two echoed stations.
        # D/E/F/G (deltas, raw bearing) stay black; H/I/J/K (final DMS +
        # distance) are blue, matching the original.
        _set(ws, f"D{r}", f"=B{from_r}-B{to_r}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"E{r}", f"=C{from_r}-C{to_r}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"F{r}", f"=ATAN2(D{r},E{r})*180/PI()", color=COMPUTED_COLOR)
        _set(ws, f"G{r}", f"=IF(F{r}<0,F{r}+360,F{r})", color=COMPUTED_COLOR)
        _set(ws, f"H{r}", f"=TRUNC(G{r},0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"I{r}", f"=TRUNC((G{r}-H{r})*60,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"J{r}", f"=ROUND((G{r}-H{r}-I{r}/60)*3600,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"K{r}", f"=SQRT(D{r}^2+E{r}^2)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        if headers:
            _set(ws, f"D{r-1}", "\u00b1\u0394N", bold=True)
            _set(ws, f"E{r-1}", "\u00b1\u0394E", bold=True)
            _set(ws, f"H{r-1}", "BEARING", bold=True)
            _merge(ws, f"H{r-1}:J{r-1}")
            _set(ws, f"K{r-1}", "DIST", bold=True)

    _set(ws, "B11", "OPENING", bold=True)
    _echo_row(12, 4)
    _echo_row(13, 6)
    _bearing_row(13, 12, 13, headers=12)

    _echo_row(15, 5)
    _echo_row(16, 13)
    _bearing_row(16, 15, 16)

    _set(ws, "B19", "CLOSING", bold=True)
    _echo_row(20, 8)
    _echo_row(21, 7)
    _bearing_row(21, 20, 21, headers=20)

    _echo_row(23, 9)
    _echo_row(24, 21)
    _bearing_row(24, 23, 24)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["F"].hidden = True  # intermediate decimal-degree calc, hidden in the original too
    return ws


# --------------------------------------------------------------------------
# TRVS sheet -- the part that resizes with the number of corner points.
# --------------------------------------------------------------------------
def build_trvs(wb, data, corners):
    ws = wb.create_sheet("TRVS")
    n = len(corners)

    headers = {"A": "Stn", "B": "Direction", "E": "Corr", "G": "Bearing",
               "L": "Distance", "N": "\u00b1\u0394N/m", "P": "Corr",
               "Q": "\u00b1\u0394E/m", "R": "Corr", "S": "Stn", "T": "N/m",
               "U": "E/m", "V": "Remarks"}
    for c, label in headers.items():
        _set(ws, f"{c}1", label, bold=True)
        _set(ws, f"{c}8", label, bold=True)
    _set(ws, "O8", "xx", bold=True)
    _merge(ws, "B1:D1"); _merge(ws, "G1:I1")
    _merge(ws, "B8:D8"); _merge(ws, "G8:I8")

    _set(ws, "D5", "TRAVERSE COMPUTATION SHEET", font=TITLE_FONT, bold=True)
    _set(ws, "A6", data.get("project", "PROJECT ------------------------------"))
    _set(ws, "A7", data.get("date_line",
         "Date-----------------Observed By-----------Compiled By-----------"))

    o = data["opening"]
    c_ = data["closing"]
    ang = data.get("angle_corrections", {})

    # ---- Opening block: rows 9, 10, 11 (fixed) -----------------------
    _set(ws, "A9", "=DTM!A12", color=DTM_LINK_COLOR)
    _set(ws, "E9", ang.get("row9", 0), font=INPUT_FONT, color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, "F9", "=G9+H9/60+I9/3600-E9/3600", color=DTM_LINK_COLOR)
    _set(ws, "G9", "=DTM!H13", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "H9", "=DTM!I13", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "I9", "=DTM!J13", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "B9", "=TRUNC(F9,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "C9", "=TRUNC((F9-B9)*60,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "D9", "=ROUND((F9-B9-C9/60)*3600,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "S9", "=A9", color=DTM_LINK_COLOR)

    _set(ws, "A10", "=DTM!A15", color=DTM_LINK_COLOR)
    _set(ws, "E10", ang.get("row10", 0), font=INPUT_FONT, color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, "F10", "=G10+H10/60+I10/3600-E10/3600", color=DTM_LINK_COLOR)
    _set(ws, "G10", "=DTM!H16", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "H10", "=DTM!I16", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "I10", "=DTM!J16", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "B10", "=TRUNC(F10,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "C10", "=TRUNC((F10-B10)*60,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "D10", "=ROUND((F10-B10-C10/60)*3600,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, "S10", "=A10", color=DTM_LINK_COLOR)

    _set(ws, "A11", "=S11", color=DTM_LINK_COLOR)
    _set(ws, "E11", "=AVERAGE(E9:E10)", color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, "S11", "=DTM!A13", color=DTM_LINK_COLOR)
    _set(ws, "T11", "=DTM!B13", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, "U11", "=DTM!C13", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)


    # ---- Corner-point block: rows 12 .. 11+n --------------------------
    first_row = 12
    last_row = 11 + n
    setup_first_row = 11          # opening fixed station
    setup_last_row = last_row + 1  # first closing row (bearing-check)

    for i, corner in enumerate(corners):
        r = first_row + i
        prev = r - 1
        _set(ws, f"S{r}", corner["name"], font=INPUT_FONT, color=COMPUTED_COLOR)
        _set(ws, f"T{r}", corner["n"], font=INPUT_FONT, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"U{r}", corner["e"], font=INPUT_FONT, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        # O displays the same value as L (the final corrected distance) --
        # weighting below is keyed off M directly instead, so this doesn't
        # create a circular reference (L <- N/Q <- P/R <- O <- L).
        _set(ws, f"O{r}", f"=L{r}", color=LEG_DIST_COLOR, number_format=NUMFMT_COORD)

        _set(ws, f"A{r}", f"=S{r}", color=None)
        if r == first_row:
            _set(ws, f"E{r}", f"=E${last_row + 6}", color=CORRECTION_COLOR, number_format=NUMFMT_INT)
        else:
            _set(ws, f"E{r}", f"=E{prev}+E${last_row + 6}", color=CORRECTION_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"F{r}", f"=G{r}+H{r}/60+I{r}/3600-E{r}/3600", color=COMPUTED_COLOR)
        _set(ws, f"J{r}", f"=IF(K{r}<0,K{r}+360,K{r})", color=COMPUTED_COLOR)
        _set(ws, f"K{r}", f"=ATAN2(N{r},Q{r})/PI()*180", color=COMPUTED_COLOR)
        _set(ws, f"G{r}", f"=TRUNC(J{r},0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"H{r}", f"=TRUNC((J{r}-G{r})*60,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"I{r}", f"=ROUND(((J{r}-G{r})-H{r}/60)*3600,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"B{r}", f"=TRUNC(F{r},0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"C{r}", f"=TRUNC((F{r}-B{r})*60,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"D{r}", f"=ROUND((F{r}-B{r}-C{r}/60)*3600,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
        _set(ws, f"L{r}", f"=ROUND((SQRT(N{r}^2+Q{r}^2)),5)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"M{r}", f"=SQRT((T{r}-T{prev})^2+(U{r}-U{prev})^2)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"N{r}", f"=T{r}-T{prev}-P{r}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"Q{r}", f"=U{r}-U{prev}-R{r}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"P{r}", f"=ROUND(($T${last_row + 4}*M{r}/$M${last_row + 2}),3)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"R{r}", f"=ROUND(($U${last_row + 4}*M{r}/$M${last_row + 2}),3)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)


    # ---- Closing block: rows last_row+1 .. last_row+3 -----------------
    cr1, cr2, cr3 = last_row + 1, last_row + 2, last_row + 3
    prev = last_row

    _set(ws, f"A{cr1}", "=DTM!A7", color=DTM_LINK_COLOR)
    _set(ws, f"E{cr1}", f"=E{prev}+E${cr3 + 3}", color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"F{cr1}", f"=G{cr1}+H{cr1}/60+I{cr1}/3600-E{cr1}/3600", color=COMPUTED_COLOR)
    _set(ws, f"J{cr1}", f"=IF(K{cr1}<0,K{cr1}+360,K{cr1})", color=COMPUTED_COLOR)
    _set(ws, f"K{cr1}", f"=ATAN2(N{cr1},Q{cr1})/PI()*180", color=COMPUTED_COLOR)
    _set(ws, f"G{cr1}", f"=TRUNC(J{cr1},0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"H{cr1}", f"=TRUNC((J{cr1}-G{cr1})*60,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"I{cr1}", f"=ROUND(((J{cr1}-G{cr1})-H{cr1}/60)*3600,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"B{cr1}", f"=TRUNC(F{cr1},0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"C{cr1}", f"=TRUNC((F{cr1}-B{cr1})*60,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"D{cr1}", f"=ROUND((F{cr1}-B{cr1}-C{cr1}/60)*3600,0)", color=COMPUTED_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"L{cr1}", f"=ROUND((SQRT(N{cr1}^2+Q{cr1}^2)),5)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"M{cr1}", f"=SQRT((T{cr1}-T{prev})^2+(U{cr1}-U{prev})^2)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"N{cr1}", f"=T{cr1}-T{prev}-P{cr1}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"Q{cr1}", f"=U{cr1}-U{prev}-R{cr1}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"P{cr1}", f"=ROUND(($T${cr3 + 1}*M{cr1}/$M${cr2}),3)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"R{cr1}", f"=ROUND(($U${cr3 + 1}*M{cr1}/$M${cr2}),3)", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"O{cr1}", f"=L{cr1}", color=LEG_DIST_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"S{cr1}", "=DTM!A7", color=DTM_LINK_COLOR)
    _set(ws, f"T{cr1}", "=DTM!B7", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"U{cr1}", "=DTM!C7", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)

    _set(ws, f"A{cr2}", "=DTM!A20", color=DTM_LINK_COLOR)
    _set(ws, f"E{cr2}", ang.get("row_close1", 0), font=INPUT_FONT, color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"F{cr2}", f"=G{cr2}+H{cr2}/60+I{cr2}/3600-E{cr2}/3600", color=DTM_LINK_COLOR)
    _set(ws, f"G{cr2}", "=DTM!H21", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"H{cr2}", "=DTM!I21", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"I{cr2}", "=DTM!J21", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"B{cr2}", f"=TRUNC(F{cr2},0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"C{cr2}", f"=TRUNC((F{cr2}-B{cr2})*60,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"D{cr2}", f"=ROUND((F{cr2}-B{cr2}-C{cr2}/60)*3600,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"L{cr2}", f"=SUM(L{first_row}:L{cr1})", bold=True, color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"M{cr2}", f"=SUM(M{first_row}:M{cr1})", bold=True, color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"N{cr2}", f"=SUM(N{first_row}:N{cr1})", bold=True, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"O{cr2}", f"=SUM(O{first_row}:O{cr1})", bold=True, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"P{cr2}", f"=SUM(P{first_row}:P{cr1})", bold=True, color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"Q{cr2}", f"=SUM(Q{first_row}:Q{cr1})", bold=True, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"R{cr2}", f"=SUM(R{first_row}:R{cr1})", bold=True, color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"S{cr2}", f"=A{cr2}", color=DTM_LINK_COLOR)
    _set(ws, f"T{cr2}", f"=T{cr1}-T11", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"U{cr2}", f"=U{cr1}-U11", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)

    _set(ws, f"A{cr3}", "=DTM!A23", color=DTM_LINK_COLOR)
    _set(ws, f"E{cr3}", ang.get("row_close2", 0), font=INPUT_FONT, color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"F{cr3}", f"=G{cr3}+H{cr3}/60+I{cr3}/3600-E{cr3}/3600", color=DTM_LINK_COLOR)
    _set(ws, f"G{cr3}", "=DTM!H24", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"H{cr3}", "=DTM!I24", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"I{cr3}", "=DTM!J24", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"B{cr3}", f"=TRUNC(F{cr3},0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"C{cr3}", f"=TRUNC((F{cr3}-B{cr3})*60,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"D{cr3}", f"=ROUND((F{cr3}-B{cr3}-C{cr3}/60)*3600,0)", color=DTM_LINK_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"S{cr3}", f"=A{cr3}", color=DTM_LINK_COLOR)
    _set(ws, f"T{cr3}", f"=N{cr2}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"U{cr3}", f"=Q{cr2}", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)

    # ---- Misclosure summary: 3 rows --------------------------------
    avg_row, text_row, final_row = cr3 + 1, cr3 + 2, cr3 + 3
    _set(ws, f"E{avg_row}", f"=AVERAGE(E{cr2}:E{cr3})", color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"T{avg_row}", f"=SQRT((Q{final_row}^2)/2-(Q{final_row}^2)/5)", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"U{avg_row}", f"=SQRT(Q{final_row}^2-T{avg_row}^2)", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)

    _set(ws, f"A{text_row}", "Angular misclsure is", color=COMPUTED_COLOR)
    _merge(ws, f"A{text_row}:D{text_row}")
    # Live label so the station count always matches the actual number of
    # points. Denominator is (number of stations x 2) -- two angle
    # observations (backsight + foresight) per station occupied.
    station_count = f"COUNTA(A11:A{cr1})"
    _set(ws, f"A{final_row}",
         f'=CONCATENATE("Angular misclosure is in ",{station_count}*2," observations or")',
         color=COMPUTED_COLOR)
    _merge(ws, f"A{final_row}:D{final_row}")
    _set(ws, f"E{final_row}", f"=E{avg_row}/({station_count}*2)", bold=True, color=CORRECTION_COLOR, number_format=NUMFMT_INT)
    _set(ws, f"G{final_row}", '" per station', bold=True, color=COMPUTED_COLOR)
    _merge(ws, f"G{final_row}:I{final_row}")

    _set(ws, f"L{final_row}", "Linear misclosure is =", color=COMPUTED_COLOR)
    _merge(ws, f"L{final_row}:P{final_row}")
    # NOTE: this is a genuinely independent figure (verified against the
    # original file) -- typically the total distance for the wider survey
    # job, not just this traverse's own perimeter -- so it stays an input.
    _set(ws, f"T{final_row}", data.get("measured_total_distance", 0), font=INPUT_FONT, bold=True, color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _merge(ws, f"T{final_row}:U{final_row}")
    _set(ws, f"Q{final_row}", f"=ROUND((M{cr2}/T{final_row}),3)", color=COMPUTED_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"R{final_row}", "m", color=COMPUTED_COLOR)
    _set(ws, f"S{final_row}", "or 1 in", bold=True, color=COMPUTED_COLOR)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["V"].width = 14
    for col in ("F", "J", "K", "M", "O"):  # intermediate/duplicate columns, hidden in the original too
        ws.column_dimensions[col].hidden = True
    return ws, {"first_row": first_row, "last_row": last_row, "cr1": cr1, "cr2": cr2, "cr3": cr3}


# --------------------------------------------------------------------------
# Field notes sheet -- reverse-derives plausible backsight/foresight field
# readings from the already-adjusted TRVS bearings, alternating occupations
# (AT a station -> backsight to the previous station -> foresight to the
# next). Generalized for any number of corner points N.
#
# Simplifications vs. the original (documented, since the original had a
# couple of internal inconsistencies that aren't worth reproducing exactly):
#   - One "index correction" (H) input per occupation, shared by every
#     reading taken from that occupation (rather than a partially-chained
#     per-block value). Physically this is more correct anyway -- index
#     error is a property of one instrument setup.
#   - The "double reading" circle constants (the small residual between a
#     face-1/face-2 pair) default to a clean 180d00'00" -- i.e. no
#     fabricated noise -- and are left as editable input cells for anyone
#     who has real raw double-face readings to enter.
#   - The angle-check (G) on the k-th reading of an occupation always
#     compares against the (k-1)-th reading of that same occupation.
# --------------------------------------------------------------------------
def build_field_notes(wb, data, corners, trvs_rows):
    ws = wb.create_sheet("Field notes")
    n = len(corners)
    first_row = trvs_rows["first_row"]
    cr1, cr2, cr3 = trvs_rows["cr1"], trvs_rows["cr2"], trvs_rows["cr3"]

    headers = {"A": "Stn", "B": "Observed", "G": "Corr", "I": "Direction",
               "M": "Final Dist", "N": "Stn", "O": "Meas Dist"}
    for c, label in headers.items():
        _set(ws, f"{c}1", label, bold=True)
        _set(ws, f"{c}7", label, bold=True)
    _merge(ws, "B1:F1"); _merge(ws, "G1:H1"); _merge(ws, "I1:K1")
    _merge(ws, "B7:F7"); _merge(ws, "G7:H7"); _merge(ws, "I7:K7")
    _set(ws, "A5", "FIELD NOTES", font=TITLE_FONT, bold=True)
    _merge(ws, "A5:O5")
    _set(ws, "M6", "=INDEX!F95")

    _set(ws, "K8", "AT")
    _set(ws, "L8", "=TRVS!A11")

    # ---- Build the ordered block list -----------------------------
    # fresh_h: this block gets its own manual "index correction" input.
    # has_g:   this block gets an angle-check formula (vs. the block named
    #          in g_h_ref, defaulting to the immediately preceding block).
    # Confirmed against two independent real files: H and G don't always
    # move together -- e.g. open2 has its own fresh H input *and* a G
    # formula comparing it to open1.
    blocks = [
        {"kind": "ref", "target": "=DTM!A12", "trvs_ref": 9, "distance": False,
         "fresh_h": True, "has_g": False},
        {"kind": "ref", "target": "=DTM!A15", "trvs_ref": 10, "distance": False,
         "fresh_h": True, "has_g": True},
    ]
    for k in range(1, n + 2):  # legs 1 .. n+1
        leg_row = first_row + (k - 1) if k <= n else cr1
        blocks.append({"kind": "fs", "target": f"=TRVS!A{leg_row}", "trvs_ref": leg_row,
                        "distance": True, "from_occ": k - 1, "to_occ": k,
                        "fresh_h": False, "has_g": True, "anchor_to_open1": (k == 1)})
        blocks.append({"kind": "bs", "from_occ": k, "fresh_h": True, "has_g": False})
    blocks.append({"kind": "ref", "target": f"=TRVS!A{cr2}", "trvs_ref": cr2, "distance": False,
                    "fresh_h": True, "has_g": False, "fresh_reset": True})
    blocks.append({"kind": "ref", "target": f"=TRVS!A{cr3}", "trvs_ref": cr3, "distance": False,
                    "fresh_h": False, "has_g": True})

    # ---- Render sequentially ---------------------------------------
    at_marker_of = {0: 8}
    foresight_arrival_base = {}
    row = 9
    prev_base = None
    open1_base = None

    for i, blk in enumerate(blocks):
        base = row
        if i == 0:
            open1_base = base

        if blk["kind"] in ("ref", "fs"):
            _set(ws, f"A{base}", blk["target"])
            _set(ws, f"I{base}", f"=TRVS!B{blk['trvs_ref']}")
            _set(ws, f"J{base}", f"=TRVS!C{blk['trvs_ref']}")
            _set(ws, f"K{base}", f"=TRVS!D{blk['trvs_ref']}")
            if blk["kind"] == "fs":
                _set(ws, f"L{base}", f"=L{at_marker_of[blk['from_occ']]}")
                _set(ws, f"M{base}", f"=TRVS!L{blk['trvs_ref']}")
                _set(ws, f"N{base}", f"=A{base}")
                _set(ws, f"O{base}", f"=M{base}/$M$6")
                foresight_arrival_base[blk["to_occ"]] = base
        else:  # backsight
            prev_at = at_marker_of[blk["from_occ"] - 1]
            _set(ws, f"A{base}", f"=L{prev_at}")
            fs_base = foresight_arrival_base[blk["from_occ"]]
            _set(ws, f"I{base}", f"=IF(I{fs_base}<180,I{fs_base}+180,I{fs_base}-180)")
            _set(ws, f"J{base}", f"=J{fs_base}")
            _set(ws, f"K{base}", f"=K{fs_base}")

        # H: fresh manual input, or a live reference to where it chains from.
        if blk["fresh_h"]:
            _set(ws, f"H{base}", 0, font=INPUT_FONT)
        elif blk.get("anchor_to_open1"):
            _set(ws, f"H{base}", f"=H{open1_base}")
        else:
            _set(ws, f"H{base}", f"=H{prev_base}")

        # G: angle check against the previous reading (or open1, for the
        # confirmed leg-1 exception). Absent entirely where has_g is False.
        if blk["has_g"]:
            if blk.get("anchor_to_open1"):
                _set(ws, f"G{base}",
                     f"=ROUND(((B{open1_base+2}+C{open1_base+2}/60+D{open1_base+2}/3600)"
                     f"-(B{base+2}+C{base+2}/60+D{base+2}/3600))/2*3600,0)")
            else:
                _set(ws, f"G{base}",
                     f"=ROUND(((B{prev_base+2}+C{prev_base+2}/60+D{prev_base+2}/3600)"
                     f"-(B{base+2}+C{base+2}/60+D{base+2}/3600))/2*3600,0)")

        if i == 0:  # very first block ever -- no G term at all
            _set(ws, f"F{base}", f"=I{base}+J{base}/60+K{base}/3600-H{base}/3600")
        else:
            _set(ws, f"F{base}", f"=I{base}+J{base}/60+K{base}/3600-(H{base}+G{base})/3600")

        _set(ws, f"B{base}", f"=TRUNC(F{base},0)")
        _set(ws, f"C{base}", f"=TRUNC((F{base}-B{base})*60,0)")
        _set(ws, f"D{base}", f"=ROUND(((F{base}-B{base})-C{base}/60)*3600,0)")
        _set(ws, f"E{base}", f"=F{base}")

        _set(ws, f"B{base+1}", f"=TRUNC(F{base+1},0)")
        _set(ws, f"C{base+1}", f"=TRUNC((F{base+1}-B{base+1})*60,0)")
        _set(ws, f"D{base+1}", f"=ROUND(((F{base+1}-B{base+1})-C{base+1}/60)*3600,0)")
        _set(ws, f"E{base+1}", f"=E{base}-(B{base+2}+C{base+2}/60+D{base+2}/3600)")
        _set(ws, f"F{base+1}", f"=IF(E{base+1}>0,E{base+1},E{base+1}+360)")

        # Double-reading residual -- clean default, editable input.
        _set(ws, f"B{base+2}", 180, font=INPUT_FONT)
        _set(ws, f"C{base+2}", 0, font=INPUT_FONT)
        _set(ws, f"D{base+2}", 0, font=INPUT_FONT)

        if blk["kind"] == "fs":
            _set(ws, f"K{base+3}", "AT")
            _set(ws, f"L{base+3}", f"=A{base}")
            at_marker_of[blk["to_occ"]] = base + 3

        prev_base = base
        # No blank spacer before close1 -- confirmed in the original files,
        # it starts immediately after the closing-check's backsight block.
        if i + 1 < len(blocks) and blocks[i + 1].get("fresh_reset"):
            row += 3
        else:
            row += 4

    end_row = prev_base + 3
    _set(ws, f"A{end_row}", "THE END", bold=True)
    _merge(ws, f"A{end_row}:O{end_row}")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["E"].hidden = True  # intermediate decimal-degree calc, hidden in the original too
    ws.column_dimensions["F"].hidden = True
    return ws


# --------------------------------------------------------------------------
# ABSTRACT sheet -- station / N / E, filtered to plot corners only.
# --------------------------------------------------------------------------
def build_abstract(wb, plot_corner_refs):
    ws = wb.create_sheet("ABSTRACT")
    _set(ws, "C3", "ABSTRACT OF RESULTS", font=TITLE_FONT, bold=True)
    _merge(ws, "C3:E3")
    _set(ws, "C4", "Stn", bold=True)
    _set(ws, "D4", "Northings", bold=True)
    _set(ws, "E4", "Eastings", bold=True)

    row = 5
    for ref in plot_corner_refs:
        _set(ws, f"C{row}", f"=TRVS!S{ref}", color=DTM_LINK_COLOR)
        _set(ws, f"D{row}", f"=TRVS!T{ref}", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"E{row}", f"=TRVS!U{ref}", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        row += 1

    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    return ws


# --------------------------------------------------------------------------
# AREA sheet -- shoelace-formula area & perimeter over the plot-corner
# subset. The polygon is closed by repeating the first point at the end.
# --------------------------------------------------------------------------
def build_area(wb, plot_corner_refs):
    ws = wb.create_sheet("AREA")
    _set(ws, "C2", "AREA COMPUTATION", font=TITLE_FONT, bold=True)
    _merge(ws, "C2:D2")
    _set(ws, "B5", "PLOT", bold=True)
    _set(ws, "B6", "POINT", bold=True); _set(ws, "C6", "N/m", bold=True)
    _set(ws, "D6", "E/m", bold=True); _set(ws, "G6", "DIST", bold=True)

    closed_refs = list(plot_corner_refs) + [plot_corner_refs[0]]  # close the polygon
    start_row = 7
    _set(ws, f"B{start_row}", f"=TRVS!S{closed_refs[0]}", color=DTM_LINK_COLOR)
    _set(ws, f"C{start_row}", f"=TRVS!T{closed_refs[0]}", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"D{start_row}", f"=TRVS!U{closed_refs[0]}", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)

    row = start_row + 1
    for ref in closed_refs[1:]:
        prev = row - 1
        _set(ws, f"B{row}", f"=TRVS!S{ref}", color=DTM_LINK_COLOR)
        _set(ws, f"C{row}", f"=TRVS!T{ref}", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"D{row}", f"=TRVS!U{ref}", color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
        _set(ws, f"E{row}", f"=D{prev}*C{row}", color=COMPUTED_COLOR)
        _set(ws, f"F{row}", f"=C{prev}*D{row}", color=COMPUTED_COLOR)
        _set(ws, f"G{row}", f"=SQRT(((C{row}-C{prev})^2)+((D{row}-D{prev})^2))", color=DTM_LINK_COLOR, number_format=NUMFMT_DIST2)
        row += 1

    last_data_row = row - 1
    sum_row = row + 1
    _set(ws, f"E{sum_row}", f"=SUM(E{start_row+1}:E{last_data_row})", bold=True, color=COMPUTED_COLOR)
    _set(ws, f"F{sum_row}", f"=SUM(F{start_row+1}:F{last_data_row})", bold=True, color=COMPUTED_COLOR)

    _set(ws, f"B{sum_row+1}", "AREA=", bold=True, color=COMPUTED_COLOR)
    _set(ws, f"C{sum_row+1}", f"=ABS((E{sum_row}-F{sum_row})/20000)", bold=True, color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"D{sum_row+1}", "HECTARES", color=COMPUTED_COLOR)
    _set(ws, f"B{sum_row+2}", "OR", bold=True, color=COMPUTED_COLOR)
    _set(ws, f"C{sum_row+2}", f"=2.471*C{sum_row+1}", bold=True, color=DTM_LINK_COLOR, number_format=NUMFMT_COORD)
    _set(ws, f"D{sum_row+2}", "ACRES", color=COMPUTED_COLOR)

    ws.column_dimensions["E"].hidden = True
    ws.column_dimensions["F"].hidden = True
    return ws


# --------------------------------------------------------------------------
def generate_workbook(data):
    """
    data: {
      "opening": {"backsight1":{"name","n","e"}, "backsight2":{...}, "start":{...}},
      "closing": {"foresight1":{...}, "foresight2":{...}, "end":{...}},
      "corners": [ {"name","n","e","is_plot_corner": bool}, ... ],
      "angle_corrections": {"row9":0, "row10":0, "row_close1":22, "row_close2":26},
      "measured_total_distance": number,   # independent total for the job; not derivable from coordinates
      "project": "optional project header text",
      "date_line": "optional date/observer/compiler line",
      "job_meta": {"job_number","client","surveyor","date","description"},   # optional, all default to ""
      "index_notes": {"instrument","district","elevation","msl_correction","scale_factor"}  # optional
    }
    Leg distances (per corner, and the closing leg) are NOT part of the
    input -- they're computed straight from each station's coordinates.
    Returns an openpyxl Workbook with 8 sheets: TOC's, Job History, INDEX,
    DTM, TRVS, Field notes, ABSTRACT, AREA.
    """
    corners = data["corners"]
    if len(corners) < 1:
        raise ValueError("Need at least one corner point")

    wb = Workbook()
    wb.remove(wb.active)

    build_toc(wb)
    build_job_history(wb, data)
    build_index_sheet(wb, data)
    build_dtm(wb, data)
    ws_trvs, rows = build_trvs(wb, data, corners)
    build_field_notes(wb, data, corners, rows)

    # Build the plot-corner row reference list in traverse order:
    # start station (row 11) -> corner rows tagged is_plot_corner -> end
    # station (first closing row, cr1), each included only if tagged.
    plot_refs = []
    if data["opening"]["start"].get("is_plot_corner", True):
        plot_refs.append(11)
    for i, corner in enumerate(corners):
        if corner.get("is_plot_corner", True):
            plot_refs.append(rows["first_row"] + i)
    if data["closing"]["end"].get("is_plot_corner", True):
        plot_refs.append(rows["cr1"])

    if len(plot_refs) < 3:
        raise ValueError("Need at least 3 plot-corner points to compute an area")

    build_abstract(wb, plot_refs)
    build_area(wb, plot_refs)

    # Match the original files' tab order exactly.
    desired_order = ["TOC's", "Job History", "INDEX", "ABSTRACT", "DTM",
                      "Field notes", "TRVS", "AREA"]
    wb._sheets = [wb[name] for name in desired_order]

    return wb


# ---------------------------------------------------------------------------
# Flask app (Vercel serverless entrypoint)
# ---------------------------------------------------------------------------
"""
Flask API for the Traverse Form generator.

POST /api/generate  with a JSON body (see traverse_generator.generate_workbook
for the schema, or backend/sample_request.json for a worked example) and get
back a ready-to-open .xlsx file with live formulas.

Run locally:
    pip install -r requirements.txt
    python app.py
    # then open frontend/index.html in a browser (or serve it separately)
"""
import io
import traceback

from flask import Flask, request, send_file, jsonify


app = Flask(__name__)

try:
    from flask_cors import CORS
    CORS(app)  # allow the static frontend (served from anywhere) to call this API
except ImportError:
    # Minimal manual CORS fallback if flask-cors isn't installed --
    # fine for local dev; for production prefer `pip install flask-cors`.
    @app.after_request
    def _add_cors_headers(resp):
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        resp.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return resp


@app.route("/api/generate", methods=["POST"])
def generate():
    if _IMPORT_ERROR is not None:
        return jsonify({"error": f"Server dependency failed to import: {_IMPORT_ERROR}"}), 500
    try:
        data = request.get_json(force=True)
        wb = generate_workbook(data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{data.get('project_filename', 'traverse_form')}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 400


@app.route("/api/health", methods=["GET"])
def health():
    if _IMPORT_ERROR is not None:
        return jsonify({"status": "degraded", "import_error": str(_IMPORT_ERROR)}), 500
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(debug=True, port=5000)
